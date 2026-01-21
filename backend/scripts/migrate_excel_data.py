#!/usr/bin/env python3
"""
Excel to Database Migration Script.

This script extracts data from the existing Excel schedule
and imports it directly into the database.

Usage:
    python scripts/migrate_excel_data.py <excel_file.xlsm>

Options:
    --dry-run    Preview changes without committing to database
    --verbose    Show detailed progress
"""

import sys
from datetime import datetime, time
from pathlib import Path
from typing import Optional

# Add the backend directory to the path
sys.path.insert(0, str(Path(__file__).parent.parent))

import argparse
from openpyxl import load_workbook

from app.database import SessionLocal
from app.models.customer import Customer, CustomerType
from app.models.driver import Driver, DriverType
from app.models.reference import Emirate, FuelBlend
from app.models.schedule import WeeklyTemplate
from app.models.tanker import DeliveryType, Tanker


def parse_time_value(cell_value) -> Optional[time]:
    """Convert Excel time to Python time object."""
    if cell_value is None:
        return None
    if isinstance(cell_value, time):
        return cell_value
    if isinstance(cell_value, datetime):
        return cell_value.time()
    try:
        # Try parsing string format
        parts = str(cell_value).split(":")
        if len(parts) >= 2:
            return time(int(parts[0]), int(parts[1]))
    except (ValueError, IndexError):
        pass
    return None


def parse_volume(vol_str) -> Optional[int]:
    """Convert volume string like '1.5K' to liters."""
    if vol_str is None:
        return None
    vol_str = str(vol_str).strip().upper()
    try:
        if vol_str.endswith("K"):
            return int(float(vol_str[:-1]) * 1000)
        return int(float(vol_str))
    except ValueError:
        return None


def normalize_customer_type(type_str: str) -> CustomerType:
    """Normalize customer type string to enum."""
    type_lower = type_str.lower().strip()
    if type_lower in ("tank", "bulk"):
        return CustomerType.BULK
    return CustomerType.MOBILE


def normalize_delivery_type(type_str: str) -> DeliveryType:
    """Normalize delivery type string to enum."""
    type_lower = type_str.lower().strip()
    if "bulk" in type_lower and "mobile" in type_lower:
        return DeliveryType.BOTH
    if "bulk" in type_lower:
        return DeliveryType.BULK
    return DeliveryType.MOBILE


class ExcelMigrator:
    """Migrates data from Excel to database."""

    def __init__(self, excel_path: str, dry_run: bool = False, verbose: bool = False):
        self.excel_path = excel_path
        self.dry_run = dry_run
        self.verbose = verbose
        self.db = SessionLocal()
        self.stats = {
            "customers_created": 0,
            "customers_skipped": 0,
            "tankers_created": 0,
            "tankers_skipped": 0,
            "drivers_created": 0,
            "drivers_skipped": 0,
            "templates_created": 0,
            "templates_skipped": 0,
        }

        # Cache for lookups
        self._emirate_cache = {}
        self._blend_cache = {}
        self._customer_cache = {}
        self._tanker_cache = {}

    def log(self, message: str, verbose_only: bool = False):
        """Print log message."""
        if verbose_only and not self.verbose:
            return
        print(message)

    def get_emirate(self, code: str) -> Optional[Emirate]:
        """Get emirate by code, with caching."""
        if code not in self._emirate_cache:
            emirate = self.db.query(Emirate).filter(Emirate.code == code).first()
            self._emirate_cache[code] = emirate
        return self._emirate_cache.get(code)

    def get_fuel_blend(self, code: str) -> Optional[FuelBlend]:
        """Get fuel blend by code, with caching."""
        if code not in self._blend_cache:
            blend = self.db.query(FuelBlend).filter(FuelBlend.code == code).first()
            self._blend_cache[code] = blend
        return self._blend_cache.get(code)

    def get_customer(self, code: str) -> Optional[Customer]:
        """Get customer by code, with caching."""
        if code not in self._customer_cache:
            customer = self.db.query(Customer).filter(Customer.code == code).first()
            self._customer_cache[code] = customer
        return self._customer_cache.get(code)

    def get_tanker(self, name: str) -> Optional[Tanker]:
        """Get tanker by name, with caching."""
        if name not in self._tanker_cache:
            tanker = self.db.query(Tanker).filter(Tanker.name == name).first()
            self._tanker_cache[name] = tanker
        return self._tanker_cache.get(name)

    def migrate_customers(self, ws):
        """Extract and migrate customer data from DataSheet."""
        self.log("\n--- Migrating Customers ---")

        # Default emirate (Dubai)
        default_emirate = self.get_emirate("DXB")

        for row in ws.iter_rows(min_row=3, max_row=40, min_col=1, max_col=5, values_only=True):
            if not row[0] or not row[1]:
                continue

            name = str(row[0]).strip()
            code = str(row[1]).strip()
            cust_type = normalize_customer_type(str(row[2]) if row[2] else "mobile")
            volume = int(row[3]) if row[3] else None
            blend_code = str(row[4]).strip() if row[4] else "B5"

            # Check if exists
            existing = self.db.query(Customer).filter(Customer.code == code).first()
            if existing:
                self.log(f"  Skipping customer (exists): {code} - {name}", verbose_only=True)
                self.stats["customers_skipped"] += 1
                self._customer_cache[code] = existing
                continue

            # Get fuel blend
            fuel_blend = self.get_fuel_blend(blend_code)

            customer = Customer(
                name=name,
                code=code,
                customer_type=cust_type,
                estimated_volume=volume,
                fuel_blend_id=fuel_blend.id if fuel_blend else None,
                emirate_id=default_emirate.id if default_emirate else None,
            )

            if not self.dry_run:
                self.db.add(customer)
                self.db.flush()
                self._customer_cache[code] = customer

            self.log(f"  Created customer: {code} - {name} ({cust_type.value})")
            self.stats["customers_created"] += 1

    def migrate_tankers(self, ws):
        """Extract and migrate tanker data from DataSheet."""
        self.log("\n--- Migrating Tankers ---")

        # Default emirate (Dubai)
        default_emirate = self.get_emirate("DXB")

        for row in ws.iter_rows(min_row=3, max_row=15, min_col=10, max_col=13, values_only=True):
            if not row[0]:
                continue

            name = str(row[0]).strip()
            capacity = int(row[1]) if row[1] else 4000
            blends_str = str(row[2]).strip() if row[2] else "B5"
            delivery_str = str(row[3]).strip() if row[3] else "both"

            # Check if exists
            existing = self.db.query(Tanker).filter(Tanker.name == name).first()
            if existing:
                self.log(f"  Skipping tanker (exists): {name}", verbose_only=True)
                self.stats["tankers_skipped"] += 1
                self._tanker_cache[name] = existing
                continue

            # Parse fuel blends
            blend_codes = [b.strip() for b in blends_str.split(",")]
            fuel_blends = [self.get_fuel_blend(code) for code in blend_codes]
            fuel_blends = [b for b in fuel_blends if b]

            # Parse delivery type
            delivery_type = normalize_delivery_type(delivery_str)
            is_3pl = "3pl" in name.lower()

            tanker = Tanker(
                name=name,
                max_capacity=capacity,
                delivery_type=delivery_type,
                is_3pl=is_3pl,
            )
            tanker.fuel_blends = fuel_blends
            if default_emirate:
                tanker.emirates = [default_emirate]

            if not self.dry_run:
                self.db.add(tanker)
                self.db.flush()
                self._tanker_cache[name] = tanker

            self.log(f"  Created tanker: {name} ({capacity}L, {delivery_type.value})")
            self.stats["tankers_created"] += 1

    def migrate_drivers(self, wb):
        """Extract and migrate driver data from various sheets."""
        self.log("\n--- Migrating Drivers ---")

        drivers_found = set()

        # Check DataSheet for driver names
        if "DataSheet" in wb.sheetnames:
            ws = wb["DataSheet"]
            for row in ws.iter_rows(min_row=3, max_row=15, min_col=19, max_col=30, values_only=True):
                for cell in row:
                    if cell and isinstance(cell, str) and not cell.isdigit():
                        name = cell.strip()
                        if name and name not in ("OFF", "HOL", "F", "") and "DRIVER" not in name.upper():
                            drivers_found.add(name)

        # Check Driver Schedule sheets
        for month in ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]:
            sheet_name = f"Driver Schedule - {month}"
            if sheet_name in wb.sheetnames:
                ws_drv = wb[sheet_name]
                for row in ws_drv.iter_rows(min_row=3, max_row=25, min_col=1, max_col=1, values_only=True):
                    if row[0] and isinstance(row[0], str):
                        name = row[0].strip()
                        if name and not name.startswith("3PL") and name not in ("OFF", "HOL", "F", ""):
                            drivers_found.add(name)

        # Create drivers
        for name in sorted(drivers_found):
            # Check if exists
            existing = self.db.query(Driver).filter(Driver.name == name).first()
            if existing:
                self.log(f"  Skipping driver (exists): {name}", verbose_only=True)
                self.stats["drivers_skipped"] += 1
                continue

            driver_type = DriverType.THIRD_PARTY if "3pl" in name.lower() else DriverType.INTERNAL

            driver = Driver(
                name=name,
                driver_type=driver_type,
            )

            if not self.dry_run:
                self.db.add(driver)

            self.log(f"  Created driver: {name} ({driver_type.value})")
            self.stats["drivers_created"] += 1

    def migrate_weekly_templates(self, wb):
        """Extract and migrate weekly template data."""
        self.log("\n--- Migrating Weekly Templates ---")

        day_names = ["Saturday", "Sunday", "Monday", "Tuesday", "Wednesday", "Thursday", "Friday"]

        for day_idx, day_name in enumerate(day_names):
            if day_name not in wb.sheetnames:
                self.log(f"  Sheet not found: {day_name}", verbose_only=True)
                continue

            ws_day = wb[day_name]

            # Find the trip table (starts with CUS_CODE header)
            trip_start_row = None
            for row_idx, row in enumerate(ws_day.iter_rows(min_row=1, max_row=30, min_col=1, max_col=1, values_only=True), 1):
                if row[0] == "CUS_CODE":
                    trip_start_row = row_idx + 1
                    break

            if not trip_start_row:
                self.log(f"  No trip table found in {day_name}", verbose_only=True)
                continue

            for row in ws_day.iter_rows(min_row=trip_start_row, max_row=trip_start_row + 40, min_col=1, max_col=8, values_only=True):
                if not row[0] or row[0] == "CUS_CODE":
                    continue

                cust_code = str(row[0]).strip()
                tanker_name = str(row[1]).strip() if row[1] else None
                start_time = parse_time_value(row[2])
                end_time = parse_time_value(row[3])
                blend_code = str(row[4]).strip() if row[4] else "B5"
                volume = parse_volume(row[5])
                is_mobile = bool(row[6]) if row[6] is not None else True
                needs_return = bool(row[7]) if row[7] is not None else False

                if not start_time or not end_time or not volume:
                    continue

                # Get customer
                customer = self.get_customer(cust_code)
                if not customer:
                    self.log(f"  Customer not found: {cust_code}", verbose_only=True)
                    self.stats["templates_skipped"] += 1
                    continue

                # Get tanker (optional)
                tanker = None
                if tanker_name and tanker_name not in ("NOT ASSIGNED", "NOT ASSINED", "--", ""):
                    tanker = self.get_tanker(tanker_name)

                # Get fuel blend
                fuel_blend = self.get_fuel_blend(blend_code)

                template = WeeklyTemplate(
                    customer_id=customer.id,
                    day_of_week=day_idx,
                    start_time=start_time,
                    end_time=end_time,
                    tanker_id=tanker.id if tanker else None,
                    fuel_blend_id=fuel_blend.id if fuel_blend else None,
                    volume=volume,
                    is_mobile_op=is_mobile,
                    needs_return=needs_return,
                )

                if not self.dry_run:
                    self.db.add(template)

                self.log(f"  Created template: {day_name} {start_time}-{end_time} -> {cust_code}", verbose_only=True)
                self.stats["templates_created"] += 1

    def run(self):
        """Run the full migration."""
        print("=" * 60)
        print("NF Dispatch Planner - Excel Data Migration")
        print("=" * 60)
        print(f"\nSource file: {self.excel_path}")
        print(f"Dry run: {self.dry_run}")
        print()

        try:
            # Load workbook
            self.log("Loading Excel file...")
            wb = load_workbook(self.excel_path, data_only=True)
            self.log(f"Found sheets: {', '.join(wb.sheetnames)}")

            # Get DataSheet
            if "DataSheet" not in wb.sheetnames:
                print("ERROR: DataSheet not found in Excel file")
                return False

            ws_data = wb["DataSheet"]

            # Run migrations
            self.migrate_customers(ws_data)
            self.migrate_tankers(ws_data)
            self.migrate_drivers(wb)
            self.migrate_weekly_templates(wb)

            # Commit or rollback
            if self.dry_run:
                self.db.rollback()
                self.log("\n[DRY RUN] Changes rolled back")
            else:
                self.db.commit()
                self.log("\nChanges committed to database")

            # Print summary
            print("\n" + "=" * 60)
            print("Migration Summary")
            print("=" * 60)
            print(f"  Customers:  {self.stats['customers_created']} created, {self.stats['customers_skipped']} skipped")
            print(f"  Tankers:    {self.stats['tankers_created']} created, {self.stats['tankers_skipped']} skipped")
            print(f"  Drivers:    {self.stats['drivers_created']} created, {self.stats['drivers_skipped']} skipped")
            print(f"  Templates:  {self.stats['templates_created']} created, {self.stats['templates_skipped']} skipped")
            print("=" * 60)

            return True

        except Exception as e:
            self.db.rollback()
            print(f"\nERROR: Migration failed - {e}")
            raise
        finally:
            self.db.close()


def main():
    parser = argparse.ArgumentParser(description="Migrate data from Excel to database")
    parser.add_argument("excel_file", help="Path to the Excel file")
    parser.add_argument("--dry-run", action="store_true", help="Preview changes without committing")
    parser.add_argument("--verbose", "-v", action="store_true", help="Show detailed progress")

    args = parser.parse_args()

    if not Path(args.excel_file).exists():
        print(f"ERROR: File not found: {args.excel_file}")
        sys.exit(1)

    migrator = ExcelMigrator(
        excel_path=args.excel_file,
        dry_run=args.dry_run,
        verbose=args.verbose,
    )

    success = migrator.run()
    sys.exit(0 if success else 1)


if __name__ == "__main__":
    main()
