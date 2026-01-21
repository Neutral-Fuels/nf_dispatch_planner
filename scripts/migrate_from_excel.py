# ============================================
# Excel to Database Migration Script
# ============================================
# This script extracts data from the existing Excel schedule
# and generates SQL insert statements for the new database.
#
# Usage: python migrate_from_excel.py Schedule_Planning_-_2025.xlsm > seed_data.sql

import sys
from openpyxl import load_workbook
from datetime import datetime, time
import json

def parse_time(cell_value):
    """Convert Excel time to string format"""
    if cell_value is None:
        return None
    if isinstance(cell_value, time):
        return cell_value.strftime('%H:%M')
    if isinstance(cell_value, datetime):
        return cell_value.strftime('%H:%M')
    return str(cell_value)

def parse_volume(vol_str):
    """Convert volume string like '1.5K' to liters"""
    if vol_str is None:
        return None
    vol_str = str(vol_str).strip().upper()
    if vol_str.endswith('K'):
        return int(float(vol_str[:-1]) * 1000)
    return int(float(vol_str))

def escape_sql(value):
    """Escape single quotes for SQL"""
    if value is None:
        return 'NULL'
    return "'" + str(value).replace("'", "''") + "'"

def main(excel_path):
    wb = load_workbook(excel_path, data_only=True)
    
    # ========================
    # Extract Reference Data
    # ========================
    print("-- ============================================")
    print("-- NF Dispatch Planner - Seed Data")
    print("-- Generated from Excel: " + excel_path)
    print("-- Generated at: " + datetime.now().isoformat())
    print("-- ============================================")
    print()
    
    # Emirates
    print("-- Emirates")
    print("INSERT INTO emirates (code, name) VALUES")
    print("    ('DXB', 'Dubai'),")
    print("    ('AUH', 'Abu Dhabi'),")
    print("    ('SHJ', 'Sharjah'),")
    print("    ('AJM', 'Ajman'),")
    print("    ('RAK', 'Ras Al Khaimah'),")
    print("    ('FUJ', 'Fujairah'),")
    print("    ('UAQ', 'Umm Al Quwain')")
    print("ON CONFLICT (code) DO NOTHING;")
    print()
    
    # Fuel Blends
    print("-- Fuel Blends")
    print("INSERT INTO fuel_blends (code, name, biodiesel_percentage) VALUES")
    print("    ('B0', 'Pure Diesel', 0),")
    print("    ('B5', '5% Biodiesel Blend', 5),")
    print("    ('B7', '7% Biodiesel Blend', 7),")
    print("    ('B10', '10% Biodiesel Blend', 10),")
    print("    ('B20', '20% Biodiesel Blend', 20),")
    print("    ('B100', 'Pure Biodiesel', 100)")
    print("ON CONFLICT (code) DO NOTHING;")
    print()
    
    # ========================
    # Extract from DataSheet
    # ========================
    ws = wb['DataSheet']
    
    # Customers
    print("-- Customers")
    customers = {}
    for row in ws.iter_rows(min_row=3, max_row=40, min_col=1, max_col=5, values_only=True):
        if row[0] and row[1]:
            name = str(row[0]).strip()
            code = str(row[1]).strip()
            cust_type = str(row[2]).strip().lower() if row[2] else 'mobile'
            volume = row[3] if row[3] else 1000
            blend = str(row[4]).strip() if row[4] else 'B5'
            
            # Normalize customer type
            if cust_type == 'tank':
                cust_type = 'bulk'
            elif cust_type not in ('bulk', 'mobile'):
                cust_type = 'mobile'
            
            # Skip duplicates
            if code not in customers:
                customers[code] = {
                    'name': name,
                    'code': code,
                    'type': cust_type,
                    'volume': volume,
                    'blend': blend
                }
    
    print("INSERT INTO customers (name, code, customer_type, estimated_volume, emirate_id, fuel_blend_id) VALUES")
    values = []
    for code, cust in customers.items():
        # Default to Dubai for emirate
        values.append(f"    ({escape_sql(cust['name'])}, {escape_sql(cust['code'])}, '{cust['type']}', {cust['volume']}, (SELECT id FROM emirates WHERE code = 'DXB'), (SELECT id FROM fuel_blends WHERE code = '{cust['blend']}'))")
    print(',\n'.join(values))
    print("ON CONFLICT (code) DO NOTHING;")
    print()
    
    # Tankers
    print("-- Tankers")
    tankers = {}
    for row in ws.iter_rows(min_row=3, max_row=15, min_col=10, max_col=13, values_only=True):
        if row[0]:
            name = str(row[0]).strip()
            capacity = int(row[1]) if row[1] else 4000
            blends = str(row[2]).strip() if row[2] else 'B5'
            delivery = str(row[3]).strip().lower() if row[3] else 'both'
            
            # Normalize delivery type
            if 'bulk' in delivery and 'mobile' in delivery:
                delivery = 'both'
            elif 'bulk' in delivery:
                delivery = 'bulk'
            else:
                delivery = 'mobile'
            
            is_3pl = '3pl' in name.lower()
            
            tankers[name] = {
                'name': name,
                'capacity': capacity,
                'blends': [b.strip() for b in blends.split(',')],
                'delivery': delivery,
                'is_3pl': is_3pl
            }
    
    print("INSERT INTO tankers (name, max_capacity, delivery_type, is_3pl) VALUES")
    values = []
    for name, tank in tankers.items():
        values.append(f"    ({escape_sql(tank['name'])}, {tank['capacity']}, '{tank['delivery']}', {str(tank['is_3pl']).lower()})")
    print(',\n'.join(values))
    print("ON CONFLICT DO NOTHING;")
    print()
    
    # Tanker-Blend relationships
    print("-- Tanker-Blend Relationships")
    for name, tank in tankers.items():
        for blend in tank['blends']:
            print(f"INSERT INTO tanker_blends (tanker_id, fuel_blend_id)")
            print(f"SELECT t.id, fb.id FROM tankers t, fuel_blends fb")
            print(f"WHERE t.name = {escape_sql(name)} AND fb.code = '{blend}'")
            print("ON CONFLICT DO NOTHING;")
    print()
    
    # Tanker-Emirates (default all to Dubai for now)
    print("-- Tanker-Emirates Relationships (default to Dubai)")
    for name in tankers.keys():
        print(f"INSERT INTO tanker_emirates (tanker_id, emirate_id)")
        print(f"SELECT t.id, e.id FROM tankers t, emirates e")
        print(f"WHERE t.name = {escape_sql(name)} AND e.code = 'DXB'")
        print("ON CONFLICT DO NOTHING;")
    print()
    
    # Drivers
    print("-- Drivers")
    drivers = set()
    # Extract driver names from DataSheet and Driver Schedule sheets
    for row in ws.iter_rows(min_row=3, max_row=15, min_col=19, max_col=30, values_only=True):
        for cell in row:
            if cell and isinstance(cell, str) and not cell.isdigit():
                drivers.add(cell.strip())
    
    # Also check Driver Schedule sheets
    for month in ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov']:
        sheet_name = f'Driver Schedule - {month}'
        if sheet_name in wb.sheetnames:
            ws_drv = wb[sheet_name]
            for row in ws_drv.iter_rows(min_row=3, max_row=25, min_col=1, max_col=1, values_only=True):
                if row[0] and isinstance(row[0], str):
                    name = row[0].strip()
                    if name and not name.startswith('3PL') and name not in ('OFF', 'HOL', 'F'):
                        drivers.add(name)
    
    # Remove special entries
    drivers = {d for d in drivers if d not in ('OFF', 'HOL', 'F', '', None) and 'DRIVER' not in d.upper()}
    
    print("INSERT INTO drivers (name, driver_type) VALUES")
    values = []
    for driver in sorted(drivers):
        if '3pl' in driver.lower():
            values.append(f"    ({escape_sql(driver)}, '3pl')")
        else:
            values.append(f"    ({escape_sql(driver)}, 'internal')")
    print(',\n'.join(values))
    print("ON CONFLICT DO NOTHING;")
    print()
    
    # ========================
    # Weekly Templates
    # ========================
    print("-- Weekly Templates")
    day_names = ['Saturday', 'Sunday', 'Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']
    
    for day_idx, day_name in enumerate(day_names):
        if day_name in wb.sheetnames:
            ws_day = wb[day_name]
            
            # Find the trip table (starts with CUS_CODE header)
            trip_start_row = None
            for row_idx, row in enumerate(ws_day.iter_rows(min_row=1, max_row=30, min_col=1, max_col=1, values_only=True), 1):
                if row[0] == 'CUS_CODE':
                    trip_start_row = row_idx + 1
                    break
            
            if trip_start_row:
                for row in ws_day.iter_rows(min_row=trip_start_row, max_row=trip_start_row+40, min_col=1, max_col=8, values_only=True):
                    if row[0] and row[0] != 'CUS_CODE':
                        cust_code = str(row[0]).strip()
                        tanker_name = str(row[1]).strip() if row[1] else None
                        start_time = parse_time(row[2])
                        end_time = parse_time(row[3])
                        blend = str(row[4]).strip() if row[4] else 'B5'
                        volume = parse_volume(row[5])
                        is_mobile = bool(row[6]) if row[6] is not None else True
                        needs_return = bool(row[7]) if row[7] is not None else False
                        
                        if start_time and end_time and volume:
                            tanker_select = 'NULL'
                            if tanker_name and tanker_name not in ('NOT ASSIGNED', 'NOT ASSINED', '--'):
                                tanker_select = f"(SELECT id FROM tankers WHERE name = {escape_sql(tanker_name)})"
                            
                            print(f"INSERT INTO weekly_templates (customer_id, day_of_week, start_time, end_time, tanker_id, fuel_blend_id, volume, is_mobile_op, needs_return)")
                            print(f"SELECT")
                            print(f"    (SELECT id FROM customers WHERE code = {escape_sql(cust_code)}),")
                            print(f"    {day_idx},")
                            print(f"    '{start_time}'::TIME,")
                            print(f"    '{end_time}'::TIME,")
                            print(f"    {tanker_select},")
                            print(f"    (SELECT id FROM fuel_blends WHERE code = '{blend}'),")
                            print(f"    {volume},")
                            print(f"    {str(is_mobile).lower()},")
                            print(f"    {str(needs_return).lower()}")
                            print(f"WHERE EXISTS (SELECT 1 FROM customers WHERE code = {escape_sql(cust_code)});")
                            print()
    
    print("-- ============================================")
    print("-- Migration Complete")
    print("-- ============================================")

if __name__ == '__main__':
    if len(sys.argv) < 2:
        print("Usage: python migrate_from_excel.py <excel_file.xlsm>", file=sys.stderr)
        sys.exit(1)
    
    main(sys.argv[1])